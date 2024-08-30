import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

class ListaAsistencia:
    def __init__(self):
        self.dataframe = pd.DataFrame()

    def leer_lista_estudiantes(self, archivo_excel):

        df = pd.read_excel(archivo_excel)
        self.dataframe = df[['Id', 'Nombre']] # Solo guarda las columnas que coincidan con 'Id' y 'Nombre', debe coincidir con los campos del excel/hojas de calculo

    def crear_archivo_asistencia(self, nombre_archivo_salida):

        if self.dataframe.empty:
            print("El DataFrame está vacío. ")
            return
        
        # Funcion lambda para ordenar por apellido
        self.dataframe = self.dataframe.sort_values(by='Nombre', key=lambda x: x.str.split().str[0])

        dias_semana = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes']
        for dia in dias_semana:
            self.dataframe[dia] = ''

        self.dataframe.to_excel(nombre_archivo_salida, index=False)

        self.estilo_excel(nombre_archivo_salida)

        print(f"Archivo {nombre_archivo_salida} creado exitosamente.")

    def estilo_excel(self, nombre_archivo):

        wb = load_workbook(nombre_archivo)
        ws = wb.active
        
        fill_id = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Color para id
        fill_nombre = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Color para nombre
        fill_dias = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Color para dias
        fill_rango_nombres_id = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid") # Color para filas de nombre e id

        ws['A1'].fill = fill_id  # Columna Identificación
        ws['B1'].fill = fill_nombre  # Columna Nombre

        for fila in range(2, ws.max_row + 1):  # Colores para los nombres e identificaciones
            ws[f'A{fila}'].fill = fill_rango_nombres_id 
            ws[f'B{fila}'].fill = fill_rango_nombres_id 

        for col in ['C', 'D', 'E', 'F', 'G']:  # Columnas para los dias de la semana
            ws[f'{col}1'].fill = fill_dias

        ws.column_dimensions['B'].width = 30  

        wb.save(nombre_archivo)

    
    def marcar_justificaciones(self, archivo_justificaciones, archivo_asistencia):

        justificaciones_df = pd.read_excel(archivo_justificaciones)
        
        wb = load_workbook(archivo_asistencia)
        ws = wb.active

        fill_justificado = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid") # Se marca con azul los dias que estan justificados de los respectivos estudiantes

        # Recorre el archivo con los estudiantes que tienen justificacion valida
        for index, row in justificaciones_df.iterrows():
            identificacion = row['Id']
            dias_justificados = {
                'Lunes': row['Lunes'],
                'Martes': row['Martes'],
                'Miércoles': row['Miércoles'],
                'Jueves': row['Jueves'],
                'Viernes': row['Viernes']
            }

            # Encontrar la fila correspondiente al estudiante en el archivo de asistencia
            for fila in range(2, ws.max_row + 1):
                if ws[f'A{fila}'].value == identificacion:
                    for dia, justificado in dias_justificados.items():
                        if justificado:  
                            columna = {
                                'Lunes': 'C',
                                'Martes': 'D',
                                'Miércoles': 'E',
                                'Jueves': 'F',
                                'Viernes': 'G'
                            }[dia]
                            ws[f'{columna}{fila}'].fill = fill_justificado

        wb.save(archivo_asistencia)

# Ejemplos de uso

#lista_asistencia = ListaAsistencia()
#lista_asistencia.leer_lista_estudiantes("estudiantes.xlsx")
#lista_asistencia.crear_archivo_asistencia("asistencia_semanal_G11-P1.xlsx") # Nota: si se tiene abierto el archivo abierto al ejecutar esta linea genera un error
#lista_asistencia.marcar_justificaciones("justificaciones.xlsx", "asistencia_semanal_G11-P1.xlsx") #Nota: si se tiene abierto el archivo abierto al ejecutar esta linea genera un error
