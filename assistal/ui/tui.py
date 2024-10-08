import assistal.config as C
import assistal.fetcher as fetcher
import assistal.ui.commons as commons
from assistal.xlsx import XLSX
from assistal.logger import plog
from assistal.classes.AssistanceEntry import AssistanceEntry
from assistal.classes.Record import Record
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import pandas as pd
import assistal.email as email
import os

import assistal.ui.menus.manage_records as manage_records_


def download_document():

    commons.print_text_ascii("Descargar Documento")
    fetcher.download_google_drive_file(C.GOOGLE_DRIVE_RECORDS_DOCUMENT, C.RUNTIME_RECORDS_FILE, merge_criteria=["timestamp", "identificacion"])
    
def manage_students():

    commons.print_text_ascii("Gestionar Estudiates")

def manage_records():

    commons.print_text_ascii("Gestionar Fichas")
    manage_records_.run()

def generate_assistance():

    commons.print_text_ascii("Generar Asistencia")

    records_doc = XLSX(C.RUNTIME_RECORDS_FILE, Record.get_fields_listed())

    if not records_doc: return
    records_doc.write_on_change = True

    assistance_doc = XLSX(C.RUNTIME_ASSISTANCE_FILE, AssistanceEntry.get_fields_listed())

    if not assistance_doc:
        plog("warning", "no se pudo leer el archivo de asistencia")

    assistance_doc.write_on_change = True

    for _, row in records_doc.df.iterrows():
        record = row.to_dict()

        new_entry = {
            "identificacion": record["identificacion"],
            "nombre_estudiante": record["nombre_estudiante"],
            "estado": "presente" if record["estado"] == "aceptado" else "ausente",
            "grado": record["grado"],
            "grupo": record["grupo"]
        }

        update_query = {
            "identificacion": record["identificacion"],
            "nombre_estudiante": record["nombre_estudiante"]
        }

        ok, level, message = True, "info", f"se pudo verificar la asistencia de {record['nombre_estudiante']}"

        ok = assistance_doc.update_entry(update_query, new_entry)
        if not ok:
            ok = assistance_doc.create_entry(new_entry)

        if not ok:
            level, message = "warning", "no " + message

        plog(level, message)

    records_doc = XLSX(C.RUNTIME_RECORDS_FILE, Record.get_fields_listed())

    accepted_records = records_doc.df[records_doc.df["estado"] == "aceptado"]
    accepted_records = accepted_records[["identificacion", "grado", "grupo", "hora", "dia"]]

    # Cada que se genera la asistencia, se actualiza el archivo de lista de asistencia semanal
    for _, row in accepted_records.iterrows():
        temp_file = C._join(C.RUNTIME_GROUPS_DIR, str(row["grado"]), f"lista_asistencia_{str(row['grado'])}-{str(row['grupo'])}.xlsx")
        
        if row["hora"] == "toda":
            horas = [1, 2, 3, 4, 5, 6]
        
        else:
            try:
                horas = row["hora"].split("-")
                horas = [i for i in range(int(horas[0]), int(horas[1]) + 1)]

            except:
                horas = [row["hora"]]
        # Leer el archivo de lista de asistencia
        wb = load_workbook(temp_file)
        ws = wb.active

        # Encontrar la fila que coincida con la identificación del estudiante
        fila_estudiante = None
        for fila in range(3, ws.max_row + 1):
            if ws[f'A{fila}'].value == row['identificacion']:
                fila_estudiante = fila
                break

        if fila_estudiante is None:
            plog("warning", f"No se encontró al estudiante con ID {row['identificacion']} en el archivo {temp_file}")
            continue

        # Determinar la columna del día
        dias_semana = ['Lunes', 'Martes', 'Miercoles', 'Jueves', 'Viernes']
        try:
            col_inicio = 3 + dias_semana.index(row["dia"].capitalize()) * 6  # Multiplicamos por 6 porque cada día tiene 6 horas
        except ValueError:
            plog("warning", f"Día {row['dia']} no válido en el registro de {row['identificacion']}")
            continue

        # Aplicar formato azul a las celdas de las horas correspondientes
        fill_azul = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Color azul claro

        for hora in horas:
            col_hora = col_inicio + hora - 1  # Ajustar columna según la hora (1-6)
            celda = ws.cell(row=fila_estudiante, column=col_hora)
            celda.fill = fill_azul

        # Guardar el archivo después de modificarlo
        wb.save(temp_file)


    print()
    assistance_doc.pretty_print()

    # Enviar correos a los grados-cursos correspondientes

    commons.print_text_ascii("Enviar correos")
    response: int = commons.show_form({"Deseas enviar los correos con las listas de asistencia semanales?": ["si", "no"]})[0]

    if response == "no":
            return
    
    if response == "si":
        emails_doc = XLSX(C._join(C.RUNTIME_DIR, "Correos_Grupos.xlsx"), ["grado", "grupo", "email"])
        
        for _, row in emails_doc.df.iterrows():
            grado, grupo, to_email = row["grado"], row["grupo"], row["email"]
            temp_file = C._join(C.RUNTIME_GROUPS_DIR, str(grado), f"lista_asistencia_{str(grado)}-{str(grupo)}.xlsx")
            email.send_email("Lista de Asistencia Semanal", "Adjunto se encuentra la lista de asistencia semanal", to_email, C.USER_EMAIL, C.USER_PASSWORD, temp_file)
            print(f"Correo enviado a {to_email} con la lista de asistencia semanal")
            plog("info", f"Correo enviado a {to_email} con la lista de asistencia semanal")


    commons.show_form({"Presiona <Enter> para regresar": str}, allow_empty=True)

def system_open_dir(path):
    system = os.name

    if os.name == "nt":
        os.startfile(path)
    elif os.name == "posix":  # macOS and Linux
        if "Darwin" in os.uname().sysname: # macos specific
            os.system(f"open '{path}'")
        else:  # linux or other POSIX systems
            os.system(f"xdg-open '{path}'")
    else:
        raise OSError("Unsupported operating system")

def open_general_assistance_directory():

    commons.print_text_ascii("Carpeta de Asistencia")
    system_open_dir(C.RUNTIME_ASSISTANCE_DIR)

def open_groups_dir():

    commons.print_text_ascii("Carpeta de Grupos")
    system_open_dir(C.RUNTIME_GROUPS_DIR)

MENU_OPTIONS = {
    "⬇️  Descargar documento con las fichas": download_document,
    "💻  Gestionar las fichas": manage_records,
    "🧒  Gestionar estudiantes": manage_students,
    "📋  Generar asistencia": generate_assistance,
    "📁  Abrir carpeta de asistencia general": open_general_assistance_directory,
    "📁  Abrir carpeta de grupos": open_groups_dir
}

def run():
    commons.show_menu(MENU_OPTIONS, "ASSISTAL", use_small_banner=False, show_exit=True)
