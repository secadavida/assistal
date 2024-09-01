import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook

class Estudiante:
    def __init__(self, nombre, estudiante_id):
        self.nombre = nombre
        self.estudiante_id = estudiante_id

    @classmethod
    def crear_archivo(cls, archivo_excel: str) -> None:
        df = pd.DataFrame(columns=['Nombre', 'Id'])

        df.to_excel(archivo_excel, index=False)
        print(f"\nArchivo {archivo_excel} creado con las columnas 'Nombre' e 'ID'.\n")

    def agregar_a_excel(self, archivo_excel):
        try:
            libro = load_workbook(archivo_excel)
            hoja = libro.active

            nuevo_dato = pd.DataFrame({
                'Nombre': [self.nombre],
                'ID': [self.estudiante_id]
            })

            datos_existentes = pd.DataFrame(hoja.values)
            datos_existentes.columns = datos_existentes.iloc[0]  
            datos_existentes = datos_existentes[1:]  

            df_final = pd.concat([datos_existentes, nuevo_dato], ignore_index=True)

            for i, col in enumerate(df_final.columns):
                for j, value in enumerate(df_final[col], start=1):
                    hoja.cell(row=j + 1, column=i + 1, value=value)

            libro.save(archivo_excel)
            print(f"Estudiante {self.nombre} con ID {self.estudiante_id} añadido al archivo {archivo_excel}.")
        
        except FileNotFoundError:
            print(f"El archivo {archivo_excel} no existe.")
        except Exception as e:
            print(f"Ocurrió un error: {e}")


# Ejemplo de uso
# estudiante = Estudiante("Juan Perez", 12345)
#
# estudiante.crear_archivo("estudiantes.xlsx")
#
# estudiante.agregar_a_excel("estudiantes.xlsx")
