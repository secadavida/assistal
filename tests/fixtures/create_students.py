from faker import Faker
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill


_TEST_ROOT_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
GROUPS_DIR = os.path.join(_TEST_ROOT_DIR, "runtime", "groups")

grades = [9, 10, 11]
groups = ['A', 'B']

num_records = 10
fake = Faker()

def create_and_write_data(file):
    data = {
        'Id': [fake.unique.random_int(min=1000, max=9999) for _ in range(num_records)],
        'Nombre': [fake.name() for _ in range(num_records)]
    }
    df = pd.DataFrame(data)
    df.to_excel(file, index=False, engine='openpyxl')

class AttendanceList():
    def __init__(self):
        self.df = pd.DataFrame()

    def ReadStudentsList(self, archivo_excel):
        """Lee un archivo Excel y almacena los nombres e identificaciones de los estudiantes."""
        df = pd.read_excel(archivo_excel)
        self.df = df[['Id', 'Nombre']]

    def CreateAttendanceFile(self, nombre_archivo_salida):
        """Crea un archivo Excel con la identificacion, nombre y columnas para los dias divididos en 6 horas, ordenados por apellido."""
        if self.df.empty:
            print("El DataFrame está vacío. Asegúrate de haber leído la lista de estudiantes primero.")
            return
        
        # Ordenar por apellido
        self.df = self.df.sort_values(by='Nombre', key=lambda x: x.str.split().str[-1])

        # Agregar columnas para los días de la semana, cada día con 6 horas
        horas = [1, 2, 3, 4, 5, 6]
        dias_semana = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes']

        columnas = [f'{dia} {hora}' for dia in dias_semana for hora in horas]
        self.df = pd.concat([self.df, pd.DataFrame(columns=columnas)])

        self.df.to_excel(nombre_archivo_salida, index=False)
        self.ApplyStyle(nombre_archivo_salida, dias_semana, horas)

    def ApplyStyle(self, nombre_archivo, dias_semana, horas):
        """Aplica colores, fusiona celdas, ajusta el formato y agrega bordes a las horas en el archivo Excel."""
        wb = load_workbook(nombre_archivo)
        ws = wb.active

        fill_id = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Color para id
        fill_nombre = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Color para nombre
        fill_dias = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Color para días
        fill_rango_nombres_id = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")  # Color para filas de nombre e id

        # Crear el borde con 4 lados
        borde = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Fusionar celdas para Id y Nombre
        ws.merge_cells('A1:A2')
        ws.merge_cells('B1:B2')
        ws['A1'] = 'Id'
        ws['B1'] = 'Nombre'
        ws['A1'].fill = fill_id
        ws['B1'].fill = fill_nombre

        # Aplicar estilo de color a las filas de nombres e id
        for fila in range(3, ws.max_row + 1):
            ws[f'A{fila}'].fill = fill_rango_nombres_id
            ws[f'B{fila}'].fill = fill_rango_nombres_id

        # Aplicar fusión y formato a los días de la semana
        col_num = 3
        for dia in dias_semana:
            # Fusionar celdas de cada día
            ws.merge_cells(start_row=1, start_column=col_num, end_row=1, end_column=col_num + 5)
            ws.cell(row=1, column=col_num).value = dia
            ws.cell(row=1, column=col_num).fill = fill_dias
            
            # Ajustar ancho de columna a 30
            for i in range(col_num, col_num + 6):
                ws.column_dimensions[ws.cell(row=2, column=i).column_letter].width = 5
            
            # Aplicar bordes y color a cada hora debajo de cada día
            for hora in horas:
                celda_hora = ws.cell(row=2, column=col_num)
                celda_hora.value = hora
                celda_hora.fill = fill_dias
                celda_hora.border = borde  # Aplicar bordes a las horas
                col_num += 1

        # Ajustar ancho de la columna Nombre
        ws.column_dimensions['B'].width = 20

        # Guardar el archivo con las modificaciones
        wb.save(nombre_archivo)
        print(f"Archivo {nombre_archivo} guardado exitosamente.")


def generate_data():
    for grade in grades:
        for group in groups:
            grade_dir = os.path.join(GROUPS_DIR, str(grade))
            group_file = os.path.join(grade_dir, f"estudiantes_{grade}-{group}.xlsx")
            asistencia_file = os.path.join(grade_dir, f"lista_asistencia_{grade}-{group}.xlsx")

            if not os.path.isdir(grade_dir):
                os.mkdir(grade_dir)

            if not os.path.isfile(group_file):
                with open(group_file, 'w'):
                    pass

            print(f"\t+ Creando estudiantes de {grade}-{group}")
            create_and_write_data(group_file)

            # Crear lista de asistencia
            print(f"\t\t+ Creando lista de asistencia para {grade}-{group}")
            lista_asistencia = AttendanceList()
            lista_asistencia.ReadStudentsList(group_file)
            lista_asistencia.CreateAttendanceFile(asistencia_file)

def run():
    if not GROUPS_DIR:
        print("Primero se necesita crear la carpeta de runtime ejecutando la aplicación")

    print("Generando estudiantes y listas de asistencia:")
    print(f"\t+ Grados: {grades}")
    print(f"\t+ Grupos: {groups}")
    print()

    generate_data()
